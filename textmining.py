__author__ = 'Williewonka'

import openpyxl

class TextMining:

    def __init__(self, mainFileName):#, rawinput):
        self.mainFileName = mainFileName
        # self.rawinput = rawinput
        self.droplist = ["", "the", "a", "an", "on", "can", "is", "not", "and", "are", "to", "in", "for", "as", "of", "it", "if", "in", "e.g", "i.e",
                         "or", "at","by", "be", "so", "with", "thus", "with", "use", "from", "such", "has", "into", "over", "via", "which", "but",
                         "that"]
        self.wb = openpyxl.load_workbook(self.mainFileName)
        self.sheet = self.wb.get_active_sheet()
        
    def Parse_Company_Counting(self,ExportName):
        companies = {}
        data = (self.sheet.columns[0], self.sheet.columns[3])
        for i in range(0, len(data[0])):
            for entry in data[1][i].value.split(";"):
                entry = entry.strip()
                company = ""
                company = entry.split("(")[1].split(")")[0].split("-")[0]
                try:
                    countrylist = companies[company]
                except:
                    countrylist = {}
                for code in data[0][i].value.split(";"):
                    code = code.strip()
                    countrycode = code[0] + code[1]
                    try:
                        countrylist[countrycode] += 1
                    except:
                        countrylist[countrycode] = 1
                companies[company] = countrylist
        exportwb = openpyxl.Workbook()
        sheet = exportwb.get_active_sheet()
        columns = []

        for countrylist in list(companies.values()):
            for country in list(countrylist.keys()):
                if country not in columns:
                    columns.append(country)

        columns = sorted(columns)
        for i in range(0,len(columns)):
            sheet.cell(row=0,column=(i+1)).value = columns[i]

        for i in range(0, len(list(companies.keys()))):
            sheet.cell(row=(i+1),column=0).value = list(companies.keys())[i]
            for country in list(list(companies.values())[i].keys()):
                sheet.cell(row=(i+1),column=(columns.index(country) + 1)).value = list(companies.values())[i][country]
        exportwb.save(ExportName)

    def Parse_Word_Counting(self, ExportName):
        data = self.sheet.columns[1]
        words = {}
        for row in data:
            #iterate through the words
            for word in row.split(' '):
                #check if word is in droplist or is an integer, in that case skip the word
                if word in self.droplist or '(' in word or ')' in word:
                    continue
                try:
                    float(word)
                    continue
                except:
                    pass
                #if word has passed checks, see if it is already in dictionary, if so add 1 if not create entry
                try:
                    words[word] += 1
                except:
                    words[word] = 1

        #export data
        #create workbook
        exportwb = openpyxl.Workbook()
        sheet = exportwb.get_active_sheet()
        #iterate through the results and write to workbook
        wordlist = list(words.keys())
        countlist = list(words.items())
        for i in range(0,len(wordlist)):
            word = wordlist[i]
            count = int(countlist[i])
            sheet.cell(column=0,row=i).value = wordlist[i]
            sheet.cell(column=1,row=i).value = int(countlist[i])

        exportwb.save(ExportName)

    def Parse_Categories(self, ExportName):
        data = self.sheet.columns[5]
        categories = {}
        for row in data:
            for entry in row.value.split(";"):
                entry = entry.strip()
                try:
                    categorie = entry.split("(")[1].split(")")[0]
                except:
                    continue
                try:
                    categories[categorie] += 1
                except:
                    categories[categorie] = 1

        #export the data
        exportwb = openpyxl.Workbook()
        exportsheet = exportwb.get_active_sheet()
        for i in range(0, len(categories)):
            exportsheet.cell(row=i,column=0).value = list(categories.keys())[i]
            exportsheet.cell(row=i,column=1).value = list(categories.values())[i]
        exportwb.save(ExportName)