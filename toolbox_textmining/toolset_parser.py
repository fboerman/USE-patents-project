__author__ = 'williewonka'

from json import loads
import openpyxl

stream = open("json/toolboxresult_toolboxes.json", "r")
inputjson = loads(stream.readlines()[0])
stream.close()

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
row = 0

toolboxes = list(map(str,sorted(map(int,list(inputjson.keys())))))
for toolbox in toolboxes:
    sheet.cell(row=row, column=0).value = toolbox
    row += 1
    patents = list(map(str,sorted(map(int,list(inputjson[toolbox].keys())))))
    for patent in patents:
        if inputjson[toolbox][patent] > 10:
            sheet.cell(row=row,column=0).value = patent
            sheet.cell(row=row,column=1).value = inputjson[toolbox][patent]
            row += 1
    row += 2
wb.save("json/toolbox_result.xlsx")