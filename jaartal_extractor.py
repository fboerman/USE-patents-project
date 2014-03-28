__author__ = 'williewonka'

import openpyxl

wb = openpyxl.load_workbook("json/jaartallen.xlsx")
sheet = wb.get_active_sheet()
wb_export = openpyxl.Workbook()
sheet_export = wb_export.get_active_sheet()

for cell in sheet.columns[1]:
    if cell.value is None:
        continue
    jaar = 2014
    # try:
    for entry in cell.value.split(" "):
        try:
            jaartal = int(entry.strip(";"))
        except:
            continue
        # jaartal = int(entry.split(" ").pop())
        if jaar > jaartal > 1900:
            jaar = jaartal
    # except:
    #     continue
    sheet_export.cell(row=cell.row,column=0).value = sheet.cell(row=cell.row,column=0).value
    sheet_export.cell(row=cell.row,column=1).value = jaar

wb_export.save("json/jaartallen_opgeschoont.xlsx")